''' I want to copy files from different folders to one folder '''
import os
import pandas as pd
from openpyxl import Workbook,load_workbook
source_dir = '/media/vlpl-host/Data/Lockbox_AB/localiser_evaluation/dataset/v2_21_07_2021/set_10_real_images/'
dest_dir = '/media/vlpl-host/Data/Lockbox_AB/localiser_evaluation/dataset/v2_21_07_2021/set_10_real_images/'
excel_output = 'set_10_quantitative_stat.xlsx'
filename = os.path.join(dest_dir, excel_output )
try:
    os.remove(filename)
except OSError:
    pass

df = pd.DataFrame()
sub_folders = os.listdir(source_dir)
for sub_fol in sub_folders:
    sub_sub_folders = os.listdir(os.path.join(source_dir, sub_fol))
    for fol in sub_sub_folders:
        if fol.endswith('.xlsx'):
            #excel_dir = os.path.join(source_dir, sub_fol, fol)
            # workbook = load_workbook(filename=excel_dir)
            # sheet = workbook.active
            # row_count = sheet.max_row
            # sheet.insert_rows(int(row_count+1))
            # workbook.save(filename=excel_dir)
    
            df = df.append(pd.read_excel(os.path.join(source_dir, sub_fol, fol)), ignore_index=True)

df.to_excel(excel_output, index=False, sheet_name='doc_combined')