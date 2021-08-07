import cv2
import os
import glob
import time
import numpy as np
from plotly import tools
import plotly.express as px
import matplotlib.pyplot as plt
from config_parser import *
import plotly.graph_objs as go
import xml.etree.ElementTree as ET
from data_structure_v3 import data_class  
from openpyxl import Workbook,load_workbook
from plotly.subplots import make_subplots
#from matplotlib_visualization import *

iou_list=[]
final_pred_count={}
final_gt_count ={}
new_metric =[] 
metric_dict={}
hw_mp_consolidated_metric= {}
IOU =[]

NEW_IOU_CALCULATION_FIELD = IOU_CALCULATION_FIELDS + ["hw","mp"]

def excel_update(dir,location,data,IOU_THR):
	workbook = load_workbook(filename=dir)
	try:
		sheet_name = str(IOU_THR)
		sheet = workbook[sheet_name]
		sheet[location] = str(data)
		workbook.save(filename=dir)
	except:
		sheet_name = str(IOU_THR)
		workbook.create_sheet(sheet_name)
		sheet = workbook[sheet_name]
		sheet[location] = str(data)
		workbook.save(filename=dir)
	
def counter(pred_count,gt_count):
	global final_gt_count
	global final_pred_count
	for i in range (len(NEW_IOU_CALCULATION_FIELD)):
		if NEW_IOU_CALCULATION_FIELD[i] in pred_count.keys():
			if NEW_IOU_CALCULATION_FIELD[i] in final_pred_count.keys():
				final_pred_count[NEW_IOU_CALCULATION_FIELD[i]] += pred_count[NEW_IOU_CALCULATION_FIELD[i]]
			else:
				final_pred_count[NEW_IOU_CALCULATION_FIELD[i]]  = pred_count[NEW_IOU_CALCULATION_FIELD[i]]

		if NEW_IOU_CALCULATION_FIELD[i] in gt_count.keys():
			if NEW_IOU_CALCULATION_FIELD[i] in final_gt_count.keys():
				final_gt_count[NEW_IOU_CALCULATION_FIELD[i]] += gt_count[NEW_IOU_CALCULATION_FIELD[i]]
			else:
				final_gt_count[NEW_IOU_CALCULATION_FIELD[i]]  = gt_count[NEW_IOU_CALCULATION_FIELD[i]]

def hw_mp_report(excel_dir,IOU_THR):
	global hw_mp_consolidated_metric
	global final_gt_count
	global final_pred_count
	fp=fn=tp=tn=precision=recall=f1_score=accuracy= 0
	field_list = ["hw","mp"]
	keys = ["metric","iou"]
	intex =0
	for field in field_list:
		tp = hw_mp_consolidated_metric[field]["metric"][0]
		fp = hw_mp_consolidated_metric[field]["metric"][1]
		tn = hw_mp_consolidated_metric[field]["metric"][2]
		fn = hw_mp_consolidated_metric[field]["metric"][3]
		# print("hw_mp_consolidated_metric[field][metric] : ",hw_mp_consolidated_metric[field]["metric"] )
		if len(hw_mp_consolidated_metric[field]["iou"]) !=0:
			avg_iou =float (sum(hw_mp_consolidated_metric[field]["iou"])/len(hw_mp_consolidated_metric[field]["iou"]))
		else:
			avg_iou = 0

		if (tp+fp)>0:
			precision = tp /(tp+fp) 
		else:
			precision = 0
		if (tp+fn)>0:
			recall = tp/(tp+fn) 
		else:
			recall = 0

		accuracy = float((tp+tn)/(tp+tn+fp+fn))

		if (precision+recall) !=0:
			f1_score = float(2*((precision*recall)/(precision+recall)))
		else:
			f1_score = 0

		fp_location = "E"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		tp_location = "F"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		fn_location = "G"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		tn_location = "H"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		iou_location = "D"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		precision_location =  "I"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		recall_location =  "J"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		accuracy_location = "K"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		f1_score_location = "L"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		Total_GT_location = "M"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		Total_pred_location = "N"+str(len(IOU_CALCULATION_FIELDS)+2+int(intex))
		intex += 1

		excel_update(excel_dir,iou_location,round(avg_iou,4),IOU_THR)
		excel_update(excel_dir,fp_location,fp,IOU_THR)
		excel_update(excel_dir,tp_location,tp,IOU_THR)
		excel_update(excel_dir,fn_location,fn,IOU_THR)
		excel_update(excel_dir,tn_location,tn,IOU_THR)
		excel_update(excel_dir,precision_location,round(precision,4),IOU_THR)
		excel_update(excel_dir,recall_location,round(recall,4),IOU_THR)
		excel_update(excel_dir,accuracy_location,round(accuracy,4),IOU_THR)
		excel_update(excel_dir,f1_score_location,round(f1_score,4),IOU_THR)

		if field in final_gt_count.keys():

			excel_update(excel_dir,Total_GT_location,final_gt_count[field],IOU_THR)
		else :
			excel_update(excel_dir,Total_GT_location,0,IOU_THR)

		if field in final_pred_count.keys():

			excel_update(excel_dir,Total_pred_location,final_pred_count[field],IOU_THR)
		else:
			excel_update(excel_dir,Total_pred_location,0,IOU_THR)
		
	final_pred_count={}
	final_gt_count ={}
	hw_mp_consolidated_metric= {}

def consolidated_metric_v2(number_images,excel_dir,IOU_THR):
	global final_gt_count
	global final_pred_count
	global iou_list
	global new_metric
	global metric_dict
	global IOU
	
	for x in range(len(IOU_CALCULATION_FIELDS)):
		fp=fn=tp=tn=precision=recall= 0
		for y in range (number_images):
			if iou_list[y][x] != 0:
				IOU.append(iou_list[y][x])
			tp = tp + int(new_metric[y][x][0])
			fp = fp + int(new_metric[y][x][1])
			tn = tn + int(new_metric[y][x][2])
			fn = fn + int(new_metric[y][x][3])
		if (tp+fp)>0:
			precision = tp /(tp+fp) 
		else:
			precision = 0
		if (tp+fn)>0:
			recall = tp/(tp+fn) 
		else:
			recall = 0
		
		if len(IOU) !=0: 
			average = float(sum(IOU)/len(IOU))
		else:
			average = 0
		accuracy = float((tp+tn)/(tp+tn+fp+fn))

		if (precision+recall) !=0:
			f1_score = float(2*((precision*recall)/(precision+recall)))
		else:
			f1_score = 0

		iou_location = "D"+str(int(x)+2)
		excel_update(excel_dir,iou_location,round(average,4),IOU_THR)
		fp_location = "E"+str(int(x)+2)
		tp_location = "F"+str(int(x)+2)
		fn_location = "G"+str(int(x)+2)
		tn_location = "H"+str(int(x)+2)
		precision_location =  "I"+str(int(x)+2)
		recall_location =  "J"+str(int(x)+2)
		excel_update(excel_dir,fp_location,fp,IOU_THR)
		excel_update(excel_dir,tp_location,tp,IOU_THR)
		excel_update(excel_dir,fn_location,fn,IOU_THR)
		excel_update(excel_dir,tn_location,tn,IOU_THR)
		excel_update(excel_dir,precision_location,round(precision,4),IOU_THR)
		excel_update(excel_dir,recall_location,round(recall,4),IOU_THR)

		accuracy_location = "K"+str(int(x)+2)
		excel_update(excel_dir,accuracy_location,round(accuracy,4),IOU_THR)
		f1_score_location = "L"+str(int(x)+2)
		excel_update(excel_dir,f1_score_location,round(f1_score,4),IOU_THR)
		Total_GT_location = "M"+str(int(x)+2)
		Total_pred_location = "N"+str(int(x)+2)

		if IOU_CALCULATION_FIELDS[x] in final_gt_count.keys():

			excel_update(excel_dir,Total_GT_location,final_gt_count[IOU_CALCULATION_FIELDS[x]],IOU_THR)
		else :
			excel_update(excel_dir,Total_GT_location,0,IOU_THR)

		if IOU_CALCULATION_FIELDS[x] in final_pred_count.keys():

			excel_update(excel_dir,Total_pred_location,final_pred_count[IOU_CALCULATION_FIELDS[x]],IOU_THR)
		else:
			excel_update(excel_dir,Total_pred_location,0,IOU_THR)
	

	hw_mp_report(excel_dir,IOU_THR)
	
	#print("metric dict :",metric_dict)

	iou_list=[]
	# final_pred_count={}
	# final_gt_count ={}
	new_metric =[] 
	metric_dict={}
	IOU =[]
	
	# workbook = load_workbook(filename=excel_dir)
	# sheet = workbook.active
	# row_count = sheet.max_row
	# sheet["A"+str(int(row_count+1))] = "."
	# sheet["A"+str(int(row_count+2))] = "."

	# workbook.save(filename=excel_dir)

def manual_calculations(predicted,gt,iou,excel_dir):
	global count 
	for i in range (len(IOU_CALCULATION_FIELDS)):
			if IOU_CALCULATION_FIELDS[i] not in predicted.keys():
				predicted[IOU_CALCULATION_FIELDS[i]]=[0]
			if IOU_CALCULATION_FIELDS[i] == "payeename":
				print("pred payeename",predicted["payeename"])
				print("gt payeename",gt["payeename"])
				excel_update(excel_dir,"P"+str(int(count+2)),predicted["payeename"])
				excel_update(excel_dir,"Q"+str(int(count+2)),gt["payeename"])

				excel_update(excel_dir,"P"+str(int(count+3)),"IOU")
				excel_update(excel_dir,"Q"+str(int(count+3)),iou[i])

				print("count : ",count)
			if IOU_CALCULATION_FIELDS[i] == "payerdetails":
				
				print("pred payerdetails",predicted["payerdetails"])
				print("gt payerdetails",gt["payerdetails"])
				excel_update(excel_dir,"R"+str(int(count+2)),predicted["payerdetails"])
				excel_update(excel_dir,"S"+str(int(count+2)),gt["payerdetails"])

				excel_update(excel_dir,"R"+str(int(count+3)),"IOU")
				excel_update(excel_dir,"S"+str(int(count+3)),iou[i])

				print("count : ",count)
			if IOU_CALCULATION_FIELDS[i] == "car":
				print("pred car",predicted["car"])
				print("gt car",gt["car"])
				excel_update(excel_dir,"T"+str(int(count+2)),predicted["car"])
				excel_update(excel_dir,"U"+str(int(count+2)),gt["car"])

				excel_update(excel_dir,"T"+str(int(count+3)),"IOU")
				excel_update(excel_dir,"U"+str(int(count+3)),iou[i])

				print("count : ",count)
			if IOU_CALCULATION_FIELDS[i] == "bankdetails":
				print("pred bankdetails",predicted["bankdetails"])
				print("gt bankdetails",gt["bankdetails"])
				excel_update(excel_dir,"V"+str(int(count+2)),predicted["bankdetails"])
				excel_update(excel_dir,"W"+str(int(count+2)),gt["bankdetails"])

				excel_update(excel_dir,"V"+str(int(count+3)),"IOU")
				excel_update(excel_dir,"W"+str(int(count+3)),iou[i])

				print("count : ",count)
			if IOU_CALCULATION_FIELDS[i] == "checknumber":
				print("pred checknumber",predicted["checknumber"])
				print("gt checknumber",gt["checknumber"])
				excel_update(excel_dir,"X"+str(int(count+2)),predicted["checknumber"])
				excel_update(excel_dir,"Y"+str(int(count+2)),gt["checknumber"])

				excel_update(excel_dir,"X"+str(int(count+3)),"IOU")
				excel_update(excel_dir,"Y"+str(int(count+3)),iou[i])

				print("count : ",count)
			count +=1

def post_processing_main(data_frame):
	global iou_list
	global new_metric
	global hw_mp_consolidated_metric

	data_frame.run_post_processing()
	iou_list.append(data_frame.Iou)
	new_metric.append(data_frame.temp_metric)
	counter(data_frame.total_detection_count,data_frame.total_gt_count)

	hw_mp_consolidated_metric=data_frame.hw_mp_consolidated_metric

#	manual_calculations(data_frame.Predicted_filtered,data_frame.Gt_Filtered,data_frame.Iou,excel_dir)

def visualize_data_v1():
	x_values=[]
	y_values = []
	for i in range(len(IOU_CALCULATION_FIELDS)):
		x_values = IOU_CALCULATION_FIELDS[i]
		temp_value = metric_dict[IOU_CALCULATION_FIELDS[i]]
		print("temp_value : ",temp_value)
		y_values.append(temp_value[0])

	trace1 = go.Bar(x_values,y_values,name='Average IOU')

	# trace1 = go.Bar(x=['giraffes', 'orangutans', 'monkeys'],y=[20, 14, 23],name='SF Zoo')
	# trace2 = go.Bar(x=['giraffes', 'orangutans', 'monkeys'],y=[12, 18, 29],name='LA Zoo')
	# trace3 = go.Scatter(x=['giraffes', 'orangutans', 'monkeys'],y=[33,20,17],name='subplots ftw')

	fig = tools.make_subplots(rows=1, cols=1, shared_xaxes=True)

	fig.append_trace(trace1, 1,1)
	# fig.append_trace(trace1, 2, 1)
	# fig.append_trace(trace2,2,1)

	fig['layout'].update(height=800, width=800,showlegend=True)
	fig.show()

