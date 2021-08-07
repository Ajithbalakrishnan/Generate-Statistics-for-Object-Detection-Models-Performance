import os
import cv2
import glob
import time
import random
import darknet
import argparse
import numpy as np
from tqdm import tqdm 
from config_parser import *
from sklearn.utils.linear_assignment_ import linear_assignment
from post_processing_v2 import *
from data_structure_v3 import data_class
from openpyxl import Workbook,load_workbook
from xlutils.copy import copy as xl_copy
import shutil

NEW_IOU_CALCULATION_FIELD = IOU_CALCULATION_FIELDS + ["hw","mp"]

def create_excel_file(dir):
    workbook = Workbook()
    sheet = workbook.active
    workbook.save(filename=dir)

def create_excel_template(dir,total_img,dpi,iou,nw_shape):
    workbook = load_workbook(filename=dir)
    workbook.create_sheet(str(iou))
    sheet = workbook[str(iou)]
    sheet["A1"] = "Total Number Images"
    sheet["B1"] = str(total_img)
    sheet["A2"] = "IOU_Threshold"
    sheet["B2"] = str(iou)
    sheet["A3"] = "DPI"
    sheet["B3"] = dpi
    sheet["A4"] = "Model"
    sheet["B4"] = YOLO_DETECTION_WEIGHT_PATH
    sheet["A5"] = "TYPE"
    sheet["B5"] = TYPE
    sheet["A6"] = "MODEL Version"
    sheet["B6"] = YOLO_MODEL_VERSION
    sheet["A7"] = "Network IO Shape"
    sheet["B7"] = str(nw_shape[0])+","+str(nw_shape[1])

    sheet["C1"] = "Fields"
    sheet["D1"] = "Avg IOU"
    sheet["E1"] = "FP"
    sheet["F1"] = "TP"
    sheet["G1"] = "FN"
    sheet["H1"] = "TN"
    sheet["I1"] = "Precision"
    sheet["J1"] = "Recall"
    sheet["k1"] = "Accuracy"
    sheet["L1"] = "F1 Score"
    sheet["M1"] = "Total GT Fields"
    sheet["N1"] = "Total Predicted Fields"
 
    for i in range(len(NEW_IOU_CALCULATION_FIELD)):
        sheet["C"+str(int(2+i))] = NEW_IOU_CALCULATION_FIELD[i]
    workbook.save(filename=dir)
 
def check_arguments_errors(args):
    assert 0 < args.thresh < 1, "Threshold should be a float between zero and one (non-inclusive)"
    if not os.path.exists(args.config_file):
        raise(ValueError("Invalid config path {}".format(os.path.abspath(args.config_file))))
    if not os.path.exists(args.weights):
        raise(ValueError("Invalid weight path {}".format(os.path.abspath(args.weights))))
    if not os.path.exists(args.data_file):
        raise(ValueError("Invalid data file path {}".format(os.path.abspath(args.data_file))))
    if args.input and not os.path.exists(args.input):
        raise(ValueError("Invalid image path {}".format(os.path.abspath(args.input))))

def check_batch_shape(images, batch_size):
    """
        Image sizes should be the same width and height
    """
    shapes = [image.shape for image in images]
    if len(set(shapes)) > 1:
        raise ValueError("Images don't have same shape")
    if len(shapes) > batch_size:
        raise ValueError("Batch size higher than number of images")
    return shapes[0]

def load_images(images_path):
    input_path_extension = images_path.split('.')[-1]
    if input_path_extension in ['jpg', 'jpeg', 'png','tif']:
        return [images_path]
    elif input_path_extension == "txt":
        with open(images_path, "r") as f:
            return f.read().splitlines()
    else:
        return glob.glob(os.path.join(images_path, "*.jpg")) + \
            glob.glob(os.path.join(images_path, "*.png")) + \
            glob.glob(os.path.join(images_path, "*.jpeg")) +\
                glob.glob(os.path.join(images_path, "*.tif"))

def prepare_batch(images, network, channels=3):
    width = darknet.network_width(network)
    height = darknet.network_height(network)

    darknet_images = []
    for image in images:
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        image_resized = cv2.resize(image_rgb, (width, height),
                                   interpolation=cv2.INTER_LINEAR)
        custom_image = image_resized.transpose(2, 0, 1)
        darknet_images.append(custom_image)

    batch_array = np.concatenate(darknet_images, axis=0)
    batch_array = np.ascontiguousarray(batch_array.flat, dtype=np.float32)/255.0
    darknet_images = batch_array.ctypes.data_as(darknet.POINTER(darknet.c_float))
    return darknet.IMAGE(width, height, channels, darknet_images)

def resize_image(img):
	color = [255,255,255]
	h, w, c = img.shape
	mindim = np.min(img.shape[:-1])
	maxdim = np.max(img.shape[:-1])
	num = maxdim - mindim
	idx = img.shape.index(mindim)
	if idx == 0:
		ypad = num
		xpad = 0
		padded_image = cv2.copyMakeBorder(img, 0, ypad, 0, 0, cv2.BORDER_CONSTANT,value=color)
	elif idx == 1:
		ypad = 0
		xpad = num
		padded_image = cv2.copyMakeBorder(img, 0, 0, 0, xpad, cv2.BORDER_CONSTANT,value=color)
	else:
		raise Exception('Either width or height is smaller than no. of channel')
	return ypad, xpad, padded_image

def preprocessing_square(image_Org):

    image = np.zeros((image_Org.shape[0],image_Org.shape[1],3))
    grayImage = cv2.cvtColor(image_Org, cv2.COLOR_BGR2GRAY)
    (thresh, BWImage) = cv2.threshold(grayImage, 127, 255, cv2.THRESH_BINARY)
    image[:,:,0] = BWImage
    image[:,:,1] = BWImage
    image[:,:,2] = BWImage
    ypad, xpad, imageMod = resize_image(image)
    cv2.imwrite('inputImage.png',imageMod)
    imageMod = cv2.imread('inputImage.png')
    os.remove('inputImage.png')
    return imageMod

def image_detection(image_path, network, class_names, class_colors, thresh):
    # Darknet doesn't accept numpy images.
    # Create one with image we reuse for each detect
    width = darknet.network_width(network)
    height = darknet.network_height(network)
    darknet_image = darknet.make_image(width, height, 3)

    image_org = cv2.imread(image_path)
    image_mod = preprocessing_square(image_org)
    image_rgb = cv2.cvtColor(image_mod, cv2.COLOR_BGR2RGB)
    image_resized = cv2.resize(image_rgb, (width, height),
                               interpolation=cv2.INTER_LINEAR)

    darknet.copy_image_from_bytes(darknet_image, image_resized.tobytes())
    detections = darknet.detect_image(network, class_names, darknet_image, thresh=thresh)
    darknet.free_image(darknet_image)
    image = darknet.draw_boxes(detections, image_resized, class_colors)
    return cv2.cvtColor(image, cv2.COLOR_BGR2RGB), detections,image_org,image_mod

def batch_detection(network, images, class_names, class_colors,
                    thresh=0.25, hier_thresh=.5, nms=.45, batch_size=4):
    image_height, image_width, _ = check_batch_shape(images, batch_size)
    darknet_images = prepare_batch(images, network)
    batch_detections = darknet.network_predict_batch(network, darknet_images, batch_size, image_width,
                                                     image_height, thresh, hier_thresh, None, 0, 0)
    batch_predictions = []
    for idx in range(batch_size):
        num = batch_detections[idx].num
        detections = batch_detections[idx].dets
        if nms:
            darknet.do_nms_obj(detections, num, len(class_names), nms)
        predictions = darknet.remove_negatives(detections, class_names, num)
        images[idx] = darknet.draw_boxes(predictions, images[idx], class_colors)
        batch_predictions.append(predictions)
    darknet.free_batch_detections(batch_detections, batch_size)
    return images, batch_predictions

def image_classification(image, network, class_names):
    width = darknet.network_width(network)
    height = darknet.network_height(network)
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    image_resized = cv2.resize(image_rgb, (width, height),
                                interpolation=cv2.INTER_LINEAR)
    darknet_image = darknet.make_image(width, height, 3)
    darknet.copy_image_from_bytes(darknet_image, image_resized.tobytes())
    detections = darknet.predict_image(network, darknet_image)
    predictions = [(name, detections[idx]) for idx, name in enumerate(class_names)]
    darknet.free_image(darknet_image)
    return sorted(predictions, key=lambda x: -x[1])

def convert2relative(image, bbox):
    """
    YOLO format use relative coordinates for annotation
    """
    x, y, w, h = bbox
    height, width, _ = image.shape
    return x/width, y/height, w/width, h/height

def save_annotations(name, image, detections, class_names):
    """
    Files saved with image_name.txt and relative coordinates
    """
    file_name = name.split(".")[:-1][0] + ".txt"
    with open(file_name, "w") as f:
        for label, confidence, bbox in detections:
            x, y, w, h = convert2relative(image, bbox)
            label = class_names.index(label)
            f.write("{} {:.4f} {:.4f} {:.4f} {:.4f} {:.4f}\n".format(label, x, y, w, h, float(confidence)))

def batch_detection_example():
    # args = parser()
    # check_arguments_errors(args)
    batch_size = YOLO_DETECTION_BATCH_SIZE
    random.seed(3)  # deterministic bbox colors
    network, class_names, class_colors = darknet.load_network(
        YOLO_DETECTION_CONFIG_PATH,
        YOLO_DETECTION_DATA_FILE_PATH,
        YOLO_DETECTION_WEIGHT_PATH,
        batch_size=batch_size
    )
    image_names = ['data/id1620887205600.png']
    images = [cv2.imread(image) for image in image_names]
    images, detections,  = batch_detection(network, images, class_names,
                                           class_colors, batch_size=batch_size,thresh = YOLO_DETECTION_THRESHOLD)
    for name, image in zip(image_names, images):
        cv2.imwrite(name.replace("data/", ""), image)
    print("YOLO_DETECTION_THRESHOLD : ",YOLO_DETECTION_THRESHOLD)
    print(detections)

def main():

    random.seed(3)  
    network, class_names, class_colors = darknet.load_network(
        YOLO_DETECTION_CONFIG_PATH,
        YOLO_DETECTION_DATA_FILE_PATH,
        YOLO_DETECTION_WEIGHT_PATH,
        batch_size=YOLO_DETECTION_BATCH_SIZE
    )

    main_folder_src = os.listdir(INPUT_DIR)
    input_folder_name = INPUT_DIR.split('/')
    print("input_folder_name : ",input_folder_name[(len(input_folder_name)-1)])
    print("main_folder_src : ",main_folder_src)

    for sub_folder_src in main_folder_src:
        print("sub_folder_src : ",sub_folder_src)
        images = []
        sub_folder_list = os.listdir(INPUT_DIR+"/" + (str(sub_folder_src)))
        input_dir_split = INPUT_DIR.split("/")
        annotation_folder_dir = "annotations_"+str(YOLO_MODEL_VERSION)+"_"+str(input_dir_split[(len(input_dir_split))-1])+"_"+str(sub_folder_src)+"_dpi_analysis_report"+"_IOU_THR_"+str(IOU_CALCULATION_THRESHOLD)
        try:
            shutil.rmtree(INPUT_DIR+"/"+str(sub_folder_src)+"/" +annotation_folder_dir)
            print ("Sucessfully Removed annotation folder")
        except OSError as e:
            print ("Done Checking annotation folder")
        for folders in sub_folder_list:
            dir_1 = str(INPUT_DIR+"/"+str(sub_folder_src)+"/"+ str(folders))
            images_new = load_images(dir_1)
            images+=images_new

        input_dir_split = INPUT_DIR.split("/")
        
        excel_dir = INPUT_DIR+"/"+str(sub_folder_src)+"/"+str(YOLO_MODEL_VERSION)+"_"+str(input_dir_split[(len(input_dir_split))-1])+"_"+str(sub_folder_src)+"_dpi_analysis_report"+"_IOU_THR_"+str(IOU_CALCULATION_THRESHOLD)+".xlsx"
     
        create_excel_file(excel_dir)
        
        annotation_folder_dir = "annotations_"+str(YOLO_MODEL_VERSION)+"_"+str(input_dir_split[(len(input_dir_split))-1])+"_"+str(sub_folder_src)+"_dpi_analysis_report"+"_IOU_THR_"+str(IOU_CALCULATION_THRESHOLD)

        path =os.path.join(INPUT_DIR+"/"+str(sub_folder_src),annotation_folder_dir)
     
        if os.path.isdir(path) == False:
            os.mkdir(path)
        
        if VARY_IOU_THRESHOLD.lower() == "yes":
            IOU_THR = [.5,.6,.7,.8,.9]
            for x in range(len(IOU_THR)):
                iou_value = float(IOU_THR[x])
                print("IOU THRESHOLD : ",iou_value)
                index = 0
                for i in tqdm (range (len(images)),desc="Loading..."):
                    if INPUT_DIR:
                        if index >= len(images):
                            break
                        image_name = images[index]
                    else:
                        image_name = input("Enter Image Path: ")
                    prev_time = time.time()
                    image, detections,image_org,image_mod = image_detection(
                        image_name, network, class_names, class_colors, YOLO_DETECTION_THRESHOLD
                        )

                    width = darknet.network_width(network)
                    height = darknet.network_height(network)

                    if YOLO_DETECTION_SAVE_LABELS:
                        save_annotations(image_name, image, detections, class_names)
                    fps = int(1/(time.time() - prev_time))

                    data_frame = data_class(Id=i,frame=image,img_path=image_name,\
                        iou_fields=IOU_CALCULATION_FIELDS,nw_detections=detections,excel_dir=excel_dir,\
                            iou_thr=iou_value,annotation_path=path,network_w_h=(width,height),Image_org=image_org,Image_mod=image_mod)
                    try:
                        post_processing_main(data_frame)
                    except Exception as e:
                        print("Error : ",e)
                        continue
                    
                    if not YOLO_DETECTION_DONT_SHOW:
                        cv2.imshow('Inference', image)
                        if cv2.waitKey() & 0xFF == ord('q'):
                            break
                    index += 1
                    
                create_excel_template(excel_dir,len(images),dpi =str(sub_folder_src),iou=iou_value,nw_shape=(width,height))
                consolidated_metric_v2(len(images),excel_dir,iou_value)
                data_frame.clean_dict()

        if VARY_IOU_THRESHOLD.lower() == "no":
            print("IOU THRESHOLD : ",IOU_CALCULATION_THRESHOLD)
            index = 0
            for i in tqdm (range (len(images)),desc="Loading..."):
                if INPUT_DIR:
                    if index >= len(images):
                        break
                    image_name = images[index]
                else:
                    image_name = input("Enter Image Path: ")
                prev_time = time.time()
                image, detections,image_org,image_mod  = image_detection(
                    image_name, network, class_names, class_colors, YOLO_DETECTION_THRESHOLD
                    )
                if YOLO_DETECTION_SAVE_LABELS:
                    save_annotations(image_name, image, detections, class_names)
                fps = int(1/(time.time() - prev_time))

                width = darknet.network_width(network)
                height = darknet.network_height(network)

                data_frame = data_class(Id=i,frame=image,img_path=image_name,\
                    iou_fields=IOU_CALCULATION_FIELDS,nw_detections=detections,excel_dir=excel_dir,\
                        iou_thr=IOU_CALCULATION_THRESHOLD,annotation_path=path,network_w_h=(width,height),Image_org=image_org,Image_mod=image_mod)
                # try:
                #     post_processing_main(data_frame)
                # except Exception as e:
                #     print("Error : ",e)
                #     continue 
                post_processing_main(data_frame)
                
                if not YOLO_DETECTION_DONT_SHOW:
                    cv2.imshow('Inference', image)
                    if cv2.waitKey() & 0xFF == ord('q'):
                        break
                index += 1
            create_excel_template(excel_dir,len(images),dpi =str(sub_folder_src),iou=IOU_CALCULATION_THRESHOLD,nw_shape=(width,height))
            consolidated_metric_v2(len(images),excel_dir,IOU_CALCULATION_THRESHOLD)
            data_frame.clean_dict()

if __name__ == "__main__":
    main()
